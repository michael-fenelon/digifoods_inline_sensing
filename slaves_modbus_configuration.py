# For Excel functions, read, write and save results.
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import copy

class Slaves_Modbus_Config():

    def __init__(self):        
        self.file_name = "sample_separation_system_modbus_config.xlsx"
        self.wb_obj = openpyxl.load_workbook(self.file_name)      # Read xlsx workbook
        self.sheet_obj = self.wb_obj.active        
        self.dict = {}
        # self.config_dict = {"slave 1":None, "slave 2":None, "slave 3":None, "slave 4":None, "slave 5":None,}
        
        self.max_num_of_rows = 0
        self.max_num_of_columns = 0
        self.max_num_of_slaves = 0
        self.slave_number = 0
        
    def get_config(self):
        self.max_num_of_columns = self.sheet_obj.max_column        # This returns the number of columns excluding the row number.
        self.max_num_of_rows = self.sheet_obj.max_row
        self.max_num_of_slaves = int(self.max_num_of_columns // 2)        # Floor division, we need integer number of slaves, not 3.5 or so.
        print("Max number of columns = ", self.max_num_of_columns)
        print("Max number of rows = ", self.max_num_of_rows)
        print("Number of slaves = ", self.max_num_of_slaves)

        # We iterate through the sheet and read each cell at (row,column) and create a dictionary keyy
        # Append the dictionary by reading each row until we reach a None.
        # Since the number of slaves start at 1 we add 1 to range().
        for slave_num in range(1, self.max_num_of_slaves + 1):
            print("Data for slave number: ", slave_num)

            # When slave_num = 1, we read col 1,2
            # When slave_num = 2, we read col 3,4
            # When slave_num = 3, we read col 5,6 ...etc.            
            # Thus, 
            #   cell_value = self.sheet_obj.cell(row = row, column = (slave_num * 2) - 1).value, 
            #   value = self.sheet_obj.cell(row = row, column = slave_num * 2).value

            # We iterate until maximum number of rows, but since each slave has different number of entries, we break out of the loop when we read a None from a cell.
            for row in range(1, self.max_num_of_rows):

                # If the cell is on None, we break 
                cell_value = self.sheet_obj.cell(row = row, column = (slave_num * 2) - 1).value 
                # print("cell_value = ", cell_value, type(cell_value))
                if cell_value is None:
                    break

                # Generate a key for the dictionary, key = "slave_" + "slave_num" + "_" + value in cell
                key = "slave_" + str(slave_num) + "_" + cell_value
                value = self.sheet_obj.cell(row = row, column = slave_num * 2).value
                self.dict[key] = value

        # print(self.dict)

        print("\n Dictionary = ")
        for key in self.dict.keys():
            print(key, self.dict[key])


if __name__ == '__main__':    
    modbus_config = Slaves_Modbus_Config()
    modbus_config.get_config()




