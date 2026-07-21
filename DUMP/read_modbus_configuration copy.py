# For Excel functions, read, write and save results.
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import copy

class Modbus_Config():

    def __init__(self):        
        self.file_name = "sample_separation_system_modbus_config.xlsx"
        self.wb_obj = openpyxl.load_workbook(self.file_name)      # Read xlsx workbook
        self.sheet_obj = self.wb_obj.active
        self.slaves = []
        self.config_dict = {}
        # self.config_dict = {"slave 1":None, "slave 2":None, "slave 3":None, "slave 4":None, "slave 5":None,}
        
        self.max_rows = 0
        self.max_columns = 0
        self.max_num_slaves = 0
        self.slave_number = 0
        
    def get_config(self):
        self.max_columns = self.sheet_obj.max_column        # This returns the number of columns including the row number.
        self.max_rows = self.sheet_obj.max_row

        self.max_num_slaves = (self.max_columns - 1)/2
        print("Number of slaves = ", self.max_num_slaves)




        # for keys in self.config_dict:
        #     print(keys)

        # self.slave_number = 0
        # while True:
        #     if self.sheet_obj['A1'] == "Name":
        #         self.slave_number = self.slave_number + 1

        #         # Eg key = "slave_1_Coil_0"
        #         key = "slave_" + str(self.slave_number) + "_"
                
        #         # Read all the rows in two columns
        #         col_key = self.slave_number
        #         col_value = col_key + 1
        #         while True:
        #             v = self.sheet_obj.cell[]

if __name__ == '__main__':    
    modbus_config = Modbus_Config()
    # print(modbus_config.wb_obj)
    # print(modbus_config.sheet_obj)
    modbus_config.get_config()
    # print(modbus_config.sheet_obj['A1'].value)
    # print(modbus_config.sheet_obj['A2'].value)
    # print(modbus_config.sheet_obj['A15'].value)
    # print(modbus_config.sheet_obj['3'])
    # print(modbus_config.sheet_obj['A1' : 'B1'].value)
    # print(modbus_config.sheet_obj.cell(row=1, column=1))

    # for row in modbus_config.sheet_obj.iter_rows(min_row=1, max_col=2, max_row=10):
    #     for cell in row:
    #         print(cell.value)

    # print("******************")
    # for col in modbus_config.sheet_obj.iter_cols(min_row=1, max_col=2, max_row=10):
    #     for cell in col:
    #         print(cell.value)
    
    # print(modbus_config.sheet_obj.cell(row=4, column=1, value=10))



